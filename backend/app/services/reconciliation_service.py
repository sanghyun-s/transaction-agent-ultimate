"""
Reconciliation Service
----------------------
1099 Pre-Reconciliation pipeline for TAU.

Provides two processing paths:
    - Rule-based: fast, deterministic, no API cost
    - Agent-based: uses Claude Agent SDK for orchestration + natural language summary

Both paths produce the same Excel output format:
    - Sheet 1: Vendor Summary (hero sheet, one row per canonical vendor)
    - Sheet 2: Transactions (full detail)
    - Sheet 3: Summary Stats

This file contains all the reconciliation logic in one place to keep TAU's
service layer clean. Follows the same pattern as file_service.py.
"""

import re
import os
import sys
import csv
import uuid
import asyncio
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from collections import defaultdict
from typing import Optional

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===========================================================================
# Configuration
# ===========================================================================

# Where generated Excels are saved (accessible via /api/reconcile/download/{id})
OUTPUT_DIR = Path(__file__).resolve().parent.parent.parent / "reconciliation_outputs"
OUTPUT_DIR.mkdir(exist_ok=True)


# ===========================================================================
# Data classes
# ===========================================================================

@dataclass
class Transaction:
    date: Optional[str]
    description: str
    amount: float
    source: str = "bank"


@dataclass
class NormalizedVendor:
    raw_name: str
    cleaned_name: str
    canonical_name: str
    entity_type: Optional[str]
    match_confidence: float
    needs_review: bool


@dataclass
class VendorSummary:
    canonical_name: str
    entity_type: Optional[str]
    total_amount: float
    transaction_count: int
    first_payment_date: Optional[str]
    last_payment_date: Optional[str]
    raw_name_variants: list[str] = field(default_factory=list)
    match_confidence: float = 1.0
    needs_review: bool = False
    review_reasons: list[str] = field(default_factory=list)


@dataclass
class ExtractionResult:
    transactions: list[Transaction]
    raw_text: str
    pages_processed: int
    extraction_method: str
    warnings: list[str]


# ===========================================================================
# Vendor normalization
# ===========================================================================

NOISE_PATTERNS = [
    r"\*[A-Z0-9]{4,}", r"#\d+", r"\d{4,}", r"\.COM\b", r"\bMKTP\b",
    r"\bPAYPAL\s*\*", r"\bSQ\s*\*", r"\bPOS\b", r"\bACH\b", r"\bDEBIT\b", r"\bCREDIT\b",
]

ENTITY_SUFFIXES = [
    "LLC", "L.L.C.", "L.L.C", "INC", "INC.", "INCORPORATED",
    "CORP", "CORP.", "CORPORATION", "CO", "CO.", "COMPANY",
    "LTD", "LTD.", "LIMITED", "LP", "L.P.", "LLP", "PC", "P.C.", "PLLC",
]


def _strip_noise(raw: str) -> str:
    s = raw.upper().strip()
    for pattern in NOISE_PATTERNS:
        s = re.sub(pattern, " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _extract_entity_type(name: str) -> tuple[str, Optional[str]]:
    name_upper = name.upper().strip()
    for suffix in ENTITY_SUFFIXES:
        pattern = rf"\b{re.escape(suffix)}\.?\s*$"
        if re.search(pattern, name_upper):
            cleaned = re.sub(pattern, "", name_upper).strip().rstrip(",").strip()
            return cleaned, suffix.replace(".", "").upper()
    return name_upper, None


def _titlecase(name: str) -> str:
    keep_upper = {"LLC", "INC", "CORP", "CO", "LTD", "LP", "LLP", "PC", "PLLC", "USA", "US"}
    return " ".join(w.upper() if w.upper() in keep_upper else w.capitalize() for w in name.split())


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()


def _find_best_match(cleaned_name: str, vendor_list: list[str], threshold: float = 0.75) -> tuple[Optional[str], float]:
    if not vendor_list:
        return None, 0.0
    best_match, best_score = None, 0.0
    for candidate in vendor_list:
        candidate_cleaned, _ = _extract_entity_type(candidate)
        score = _similarity(cleaned_name, candidate_cleaned)
        if score > best_score:
            best_score, best_match = score, candidate
    return (best_match, best_score) if best_score >= threshold else (None, best_score)


def normalize_vendor(
    raw_name: str,
    vendor_list: Optional[list[str]] = None,
    match_threshold: float = 0.75,
    review_threshold: float = 0.85,
) -> NormalizedVendor:
    vendor_list = vendor_list or []
    cleaned = _strip_noise(raw_name)
    name_without_suffix, entity_type = _extract_entity_type(cleaned)
    match, confidence = _find_best_match(name_without_suffix, vendor_list, match_threshold)

    if match:
        canonical = match
    else:
        canonical = _titlecase(name_without_suffix)
        confidence = 1.0 if not vendor_list else confidence

    needs_review = confidence < review_threshold and bool(vendor_list)

    if len(name_without_suffix.strip()) < 3 or name_without_suffix.strip().isdigit():
        needs_review = True
        canonical = f"UNRESOLVED: {raw_name}"

    return NormalizedVendor(
        raw_name=raw_name, cleaned_name=cleaned, canonical_name=canonical,
        entity_type=entity_type, match_confidence=round(confidence, 2), needs_review=needs_review,
    )


# ===========================================================================
# PDF extraction
# ===========================================================================

TXN_LINE_PATTERNS = [
    re.compile(r"^\s*(?P<date>\d{1,2}/\d{1,2}/\d{2,4})\s+(?P<description>.+?)\s+\$?(?P<amount>-?[\d,]+\.\d{2})\s*$"),
    re.compile(r"^\s*(?P<date>\d{1,2}/\d{1,2})\s+(?P<description>.+?)\s+\$?(?P<amount>-?[\d,]+\.\d{2})\s*$"),
    re.compile(r"^\s*(?P<date>\d{4}-\d{1,2}-\d{1,2})\s+(?P<description>.+?)\s+\$?(?P<amount>-?[\d,]+\.\d{2})\s*$"),
]


def _parse_amount(amt_str: str) -> Optional[float]:
    try:
        return float(amt_str.replace(",", "").replace("$", "").strip())
    except (ValueError, AttributeError):
        return None


def _extract_from_text_lines(text: str, source: str = "bank") -> list[Transaction]:
    transactions = []
    for line in text.split("\n"):
        line = line.rstrip()
        if not line.strip():
            continue
        for pattern in TXN_LINE_PATTERNS:
            m = pattern.match(line)
            if m:
                amt = _parse_amount(m.group("amount"))
                if amt is None or amt == 0:
                    continue
                transactions.append(Transaction(
                    date=m.group("date"),
                    description=m.group("description").strip(),
                    amount=abs(amt), source=source,
                ))
                break
    return transactions


def _extract_from_tables(pdf_path: str, source: str = "bank") -> list[Transaction]:
    transactions = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table or len(table) < 2:
                    continue
                header = [str(c or "").strip().lower() for c in table[0]]
                date_idx = next((i for i, h in enumerate(header) if "date" in h), None)
                desc_idx = next((i for i, h in enumerate(header) if any(
                    k in h for k in ["description", "payee", "vendor", "merchant", "transaction"]
                )), None)
                amt_idx = next((i for i, h in enumerate(header) if any(
                    k in h for k in ["amount", "debit", "withdrawal", "payment"]
                )), None)
                if date_idx is None or desc_idx is None or amt_idx is None:
                    continue
                for row in table[1:]:
                    if not row or len(row) <= max(date_idx, desc_idx, amt_idx):
                        continue
                    date_val = str(row[date_idx] or "").strip()
                    desc_val = str(row[desc_idx] or "").strip()
                    amt_val = _parse_amount(str(row[amt_idx] or ""))
                    if date_val and desc_val and amt_val and amt_val != 0:
                        transactions.append(Transaction(
                            date=date_val, description=desc_val,
                            amount=abs(amt_val), source=source,
                        ))
    return transactions


def extract_transactions(pdf_path: str, source: str = "bank") -> ExtractionResult:
    pdf_path = str(pdf_path)
    if not Path(pdf_path).exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    warnings = []
    raw_text_parts = []
    pages_processed = 0

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages_processed += 1
            raw_text_parts.append(page.extract_text() or "")

    raw_text = "\n".join(raw_text_parts)

    table_txns = []
    try:
        table_txns = _extract_from_tables(pdf_path, source)
    except Exception as e:
        warnings.append(f"Table extraction failed: {e}")

    regex_txns = _extract_from_text_lines(raw_text, source)

    if len(table_txns) >= len(regex_txns) and table_txns:
        transactions, method = table_txns, "table"
    elif regex_txns:
        transactions, method = regex_txns, "regex"
    else:
        transactions, method = [], "none"
        warnings.append("No transactions extracted. PDF may be image-based or use a non-standard layout.")

    return ExtractionResult(
        transactions=transactions, raw_text=raw_text,
        pages_processed=pages_processed, extraction_method=method, warnings=warnings,
    )


# ===========================================================================
# Aggregation
# ===========================================================================

def aggregate_by_vendor(
    transactions: list[Transaction],
    normalized: list[NormalizedVendor],
) -> list[VendorSummary]:
    if len(transactions) != len(normalized):
        raise ValueError("transactions and normalized must have the same length")

    groups: dict[str, list[tuple[Transaction, NormalizedVendor]]] = defaultdict(list)
    for txn, norm in zip(transactions, normalized):
        groups[norm.canonical_name].append((txn, norm))

    summaries = []
    for canonical, items in groups.items():
        txns = [t for t, _ in items]
        norms = [n for _, n in items]
        dates = sorted([t.date for t in txns if t.date])
        raw_variants = list({n.raw_name for n in norms})
        min_confidence = min(n.match_confidence for n in norms)
        any_needs_review = any(n.needs_review for n in norms)
        entity_type = next((n.entity_type for n in norms if n.entity_type), None)

        review_reasons = []
        if any_needs_review:
            review_reasons.append("Low vendor name match confidence")
        if len(raw_variants) > 1 and min_confidence < 0.9:
            review_reasons.append(f"Multiple raw name variants grouped together ({len(raw_variants)})")

        summaries.append(VendorSummary(
            canonical_name=canonical, entity_type=entity_type,
            total_amount=round(sum(t.amount for t in txns), 2),
            transaction_count=len(txns),
            first_payment_date=dates[0] if dates else None,
            last_payment_date=dates[-1] if dates else None,
            raw_name_variants=raw_variants,
            match_confidence=round(min_confidence, 2),
            needs_review=any_needs_review, review_reasons=review_reasons,
        ))

    summaries.sort(key=lambda v: v.total_amount, reverse=True)
    return summaries


# ===========================================================================
# Excel generation
# ===========================================================================

HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
REVIEW_FILL = PatternFill("solid", start_color="FFF4CC")
BODY_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)


def _classify_category(summary: VendorSummary) -> str:
    name_upper = summary.canonical_name.upper()
    utility_keywords = ["VERIZON", "COMCAST", "PG&E", "AT&T", "ELECTRIC", "GAS", "WATER"]
    retail_keywords = ["AMAZON", "HOME DEPOT", "STAPLES", "OFFICE DEPOT", "WALMART", "TARGET", "COSTCO"]
    if any(k in name_upper for k in utility_keywords):
        return "Utility"
    if any(k in name_upper for k in retail_keywords):
        return "Supplies/Retail"
    if summary.entity_type in ("LLC", "INC", "CORP"):
        return "Contractor"
    if "UNRESOLVED" in name_upper:
        return "Unclear"
    return "Unclear"


VENDOR_SUMMARY_COLUMNS = [
    ("Vendor Name", 22, "left"), ("Entity Type", 12, "center"),
    ("Total Paid ($)", 15, "right"), ("# Payments", 11, "center"),
    ("First Payment", 13, "center"), ("Last Payment", 13, "center"),
    ("1099 Eligible", 14, "center"), ("W-9 on File", 12, "center"),
    ("Category", 14, "center"), ("Confidence", 11, "center"),
    ("Review Needed", 14, "center"), ("Review Reason", 40, "left"),
    ("Raw Name Variants", 40, "left"),
]


def generate_excel_report(
    output_path: str,
    transactions: list[Transaction],
    normalized: list[NormalizedVendor],
    summaries: list[VendorSummary],
):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Vendor Summary
    ws = wb.create_sheet("Vendor Summary", 0)
    ws["A1"] = "1099 PRE-RECONCILIATION WORKSHEET"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    ws.merge_cells("A1:M1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A2:M2")

    for col_idx, (label, _, _) in enumerate(VENDOR_SUMMARY_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 22

    for col_idx, (_, width, _) in enumerate(VENDOR_SUMMARY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 5
    for s in summaries:
        row_data = [
            s.canonical_name, s.entity_type or "Individual?",
            s.total_amount, s.transaction_count,
            s.first_payment_date or "", s.last_payment_date or "",
            "TBD", "TBD", _classify_category(s),
            s.match_confidence, "YES" if s.needs_review else "NO",
            "; ".join(s.review_reasons), "; ".join(s.raw_name_variants),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
            _, _, align = VENDOR_SUMMARY_COLUMNS[col_idx - 1]
            cell.alignment = {"left": LEFT, "right": RIGHT, "center": CENTER}[align]
        ws.cell(row=row, column=3).number_format = '$#,##0.00;($#,##0.00);-'
        ws.cell(row=row, column=10).number_format = '0%'
        if s.needs_review:
            for col_idx in range(1, len(VENDOR_SUMMARY_COLUMNS) + 1):
                ws.cell(row=row, column=col_idx).fill = REVIEW_FILL
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{row - 1})")
    ws.cell(row=total_row, column=3).number_format = '$#,##0.00'
    ws.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=total_row, column=4, value=f"=SUM(D5:D{row - 1})")
    ws.cell(row=total_row, column=4).font = Font(name="Arial", bold=True, size=11)
    ws.freeze_panes = "A5"

    # Sheet 2: Transactions
    tx_ws = wb.create_sheet("Transactions")
    tx_ws["A1"] = "TRANSACTION DETAIL"
    tx_ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    tx_ws.merge_cells("A1:H1")
    headers = [("Date", 12), ("Raw Description", 35), ("Canonical Vendor", 25),
               ("Entity Type", 12), ("Amount ($)", 14), ("Source", 12),
               ("Confidence", 12), ("Review?", 10)]
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = tx_ws.cell(row=3, column=col_idx, value=label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER
        tx_ws.column_dimensions[get_column_letter(col_idx)].width = width
    tx_ws.row_dimensions[3].height = 22

    for i, (txn, norm) in enumerate(zip(transactions, normalized), start=4):
        values = [txn.date or "", txn.description, norm.canonical_name,
                  norm.entity_type or "", txn.amount, txn.source,
                  norm.match_confidence, "YES" if norm.needs_review else ""]
        for col_idx, value in enumerate(values, start=1):
            cell = tx_ws.cell(row=i, column=col_idx, value=value)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
        tx_ws.cell(row=i, column=5).number_format = '$#,##0.00'
        tx_ws.cell(row=i, column=7).number_format = '0%'
        if norm.needs_review:
            for col_idx in range(1, len(headers) + 1):
                tx_ws.cell(row=i, column=col_idx).fill = REVIEW_FILL
    tx_ws.freeze_panes = "A4"

    # Sheet 3: Summary Stats
    st_ws = wb.create_sheet("Summary Stats")
    st_ws["A1"] = "RECONCILIATION SUMMARY"
    st_ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    st_ws.merge_cells("A1:B1")

    total_amount = sum(s.total_amount for s in summaries)
    vendors_over_600 = sum(1 for s in summaries if s.total_amount >= 600)
    review_needed = sum(1 for s in summaries if s.needs_review)
    llc_vendors = sum(1 for s in summaries if s.entity_type in ("LLC", "INC", "CORP"))

    stats = [
        ("", ""), ("PROCESSING METRICS", ""),
        ("Total transactions processed", len(transactions)),
        ("Unique vendors identified", len(summaries)),
        ("Total $ reconciled", total_amount),
        ("", ""), ("1099 PRE-SCREEN", ""),
        ("Vendors crossing $600 threshold", vendors_over_600),
        ("Vendors with entity suffix (LLC/Inc/Corp)", llc_vendors),
        ("", ""), ("REVIEW FLAGS", ""),
        ("Vendors needing human review", review_needed),
        ("Review rate", f"{review_needed / max(len(summaries), 1):.1%}"),
    ]
    for idx, (label, value) in enumerate(stats, start=3):
        cell_a = st_ws.cell(row=idx, column=1, value=label)
        cell_b = st_ws.cell(row=idx, column=2, value=value)
        cell_a.font = BODY_FONT; cell_b.font = BODY_FONT
        if label and not value:
            cell_a.font = Font(name="Arial", bold=True, size=11, color="1F3A5F")
        if label == "Total $ reconciled":
            cell_b.number_format = '$#,##0.00'
    st_ws.column_dimensions["A"].width = 45
    st_ws.column_dimensions["B"].width = 18

    wb.save(output_path)


# ===========================================================================
# Helpers
# ===========================================================================

def load_vendor_csv(csv_path: str) -> list[str]:
    if not csv_path:
        return []
    vendors = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        first_row = next(reader, None)
        if first_row and first_row[0].lower() not in ("vendor", "vendor name", "name", "canonical name"):
            vendors.append(first_row[0].strip())
        for row in reader:
            if row and row[0].strip():
                vendors.append(row[0].strip())
    return vendors


# ===========================================================================
# Rule-based pipeline entry point
# ===========================================================================

def run_rule_based_pipeline(
    pdf_path: str,
    vendor_csv_path: Optional[str] = None,
) -> dict:
    """Execute the full rule-based pipeline. Returns a summary dict."""
    extraction = extract_transactions(pdf_path)
    if not extraction.transactions:
        return {
            "success": False,
            "error": "No transactions could be extracted from the PDF",
            "warnings": extraction.warnings,
        }

    known_vendors = load_vendor_csv(vendor_csv_path) if vendor_csv_path else []
    normalized = [normalize_vendor(t.description, known_vendors) for t in extraction.transactions]
    summaries = aggregate_by_vendor(extraction.transactions, normalized)

    file_id = f"rulebased_{uuid.uuid4().hex}.xlsx"
    output_path = OUTPUT_DIR / file_id
    generate_excel_report(str(output_path), extraction.transactions, normalized, summaries)

    total_amount = sum(s.total_amount for s in summaries)
    vendors_over_600 = sum(1 for s in summaries if s.total_amount >= 600)
    vendors_needing_review = sum(1 for s in summaries if s.needs_review)

    return {
        "success": True,
        "mode": "rule-based",
        "file_id": file_id,
        "transaction_count": len(extraction.transactions),
        "vendor_count": len(summaries),
        "total_amount": total_amount,
        "vendors_over_600": vendors_over_600,
        "vendors_needing_review": vendors_needing_review,
        "extraction_method": extraction.extraction_method,
        "warnings": extraction.warnings,
        "vendor_preview": [_summary_to_dict(s) for s in summaries],
        "agent_summary": None,
        "agent_cost_usd": 0.0,
        "agent_tool_calls": 0,
    }


def _summary_to_dict(s: VendorSummary) -> dict:
    return {
        "name": s.canonical_name,
        "entity": s.entity_type or "Individual?",
        "total": s.total_amount,
        "count": s.transaction_count,
        "first_payment": s.first_payment_date or "",
        "last_payment": s.last_payment_date or "",
        "category": _classify_category(s),
        "confidence": s.match_confidence,
        "review": s.needs_review,
        "review_reason": "; ".join(s.review_reasons),
    }


# ===========================================================================
# Agent-based pipeline (Claude Agent SDK)
# ===========================================================================

_agent_session_state: dict = {}


def _build_agent_tools():
    """Build the Claude Agent SDK tools (lazy import to avoid breaking if SDK not installed)."""
    from claude_agent_sdk import tool, create_sdk_mcp_server

    @tool(
        "extract_pdf_transactions",
        "Extract transactions from a bank or credit card statement PDF. First step when given a PDF.",
        {"pdf_path": str}
    )
    async def extract_tool(args):
        try:
            result = extract_transactions(args["pdf_path"])
        except FileNotFoundError:
            return {"content": [{"type": "text", "text": f"Error: PDF not found at {args['pdf_path']}"}]}
        _agent_session_state["transactions"] = result.transactions
        return {"content": [{"type": "text",
            "text": f"Extracted {len(result.transactions)} transactions from the PDF "
                    f"(method: {result.extraction_method}, pages: {result.pages_processed})."
        }]}

    @tool(
        "load_vendor_list",
        "Load a CSV of known canonical vendor names (optional, improves matching).",
        {"csv_path": str}
    )
    async def load_tool(args):
        try:
            vendors = load_vendor_csv(args["csv_path"])
        except FileNotFoundError:
            return {"content": [{"type": "text", "text": f"Error: CSV not found at {args['csv_path']}"}]}
        _agent_session_state["known_vendors"] = vendors
        return {"content": [{"type": "text", "text": f"Loaded {len(vendors)} known vendors."}]}

    @tool(
        "normalize_vendors",
        "Normalize vendor names. Must be called after extract_pdf_transactions.",
        {}
    )
    async def norm_tool(args):
        if "transactions" not in _agent_session_state:
            return {"content": [{"type": "text", "text": "Error: call extract_pdf_transactions first."}]}
        known = _agent_session_state.get("known_vendors", [])
        normalized = [normalize_vendor(t.description, known) for t in _agent_session_state["transactions"]]
        _agent_session_state["normalized"] = normalized
        review_count = sum(1 for n in normalized if n.needs_review)
        return {"content": [{"type": "text",
            "text": f"Normalized {len(normalized)} vendor names. {review_count} flagged for review."
        }]}

    @tool(
        "aggregate_by_vendor",
        "Group transactions by canonical vendor and sum amounts. After normalize_vendors.",
        {}
    )
    async def agg_tool(args):
        if "normalized" not in _agent_session_state:
            return {"content": [{"type": "text", "text": "Error: call normalize_vendors first."}]}
        summaries = aggregate_by_vendor(
            _agent_session_state["transactions"], _agent_session_state["normalized"]
        )
        _agent_session_state["summaries"] = summaries
        total = sum(s.total_amount for s in summaries)
        over_600 = sum(1 for s in summaries if s.total_amount >= 600)
        return {"content": [{"type": "text",
            "text": f"Aggregated into {len(summaries)} vendors. Total: ${total:,.2f}. "
                    f"Over $600: {over_600}."
        }]}

    @tool(
        "generate_excel_report",
        "Generate the final Excel workbook. Final step after aggregate_by_vendor.",
        {"output_path": str}
    )
    async def gen_tool(args):
        if "summaries" not in _agent_session_state:
            return {"content": [{"type": "text", "text": "Error: call aggregate_by_vendor first."}]}
        output_path = args["output_path"]
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        generate_excel_report(
            output_path,
            _agent_session_state["transactions"],
            _agent_session_state["normalized"],
            _agent_session_state["summaries"],
        )
        return {"content": [{"type": "text", "text": f"Excel generated at {output_path}."}]}

    server = create_sdk_mcp_server(
        name="reconciliation-tools", version="1.0.0",
        tools=[extract_tool, load_tool, norm_tool, agg_tool, gen_tool],
    )
    return server


SYSTEM_PROMPT = """You are an expert tax preparation assistant specializing in 1099 \
pre-reconciliation work. Your job is to help accountants process bank statements to \
identify vendors who may need 1099 filings.

When given a reconciliation task:
1. Call tools in order: extract → (optional) load_vendors → normalize → aggregate → generate
2. At the end, provide a concise summary covering:
   - Total dollars reconciled
   - Number of vendors crossing the $600 1099 threshold
   - Any vendors flagged for human review (and why)

Be concise and professional. An accountant wants clear numbers and clear next steps."""


async def run_agent_pipeline(
    pdf_path: str,
    vendor_csv_path: Optional[str] = None,
    model: str = "claude-haiku-4-5-20251001",
) -> dict:
    """Execute the Claude-Agent-SDK-orchestrated pipeline."""
    global _agent_session_state
    _agent_session_state = {}

    if not os.getenv("ANTHROPIC_API_KEY"):
        return {
            "success": False,
            "error": "ANTHROPIC_API_KEY not set. Check your backend .env file.",
        }

    from claude_agent_sdk import (
        ClaudeAgentOptions, ClaudeSDKClient,
        AssistantMessage, TextBlock, ResultMessage,
    )

    file_id = f"agent_{uuid.uuid4().hex}.xlsx"
    output_path = OUTPUT_DIR / file_id

    task = f"""Please process this reconciliation task:

PDF statement: {pdf_path}
Output Excel path: {output_path}"""
    if vendor_csv_path:
        task += f"\nKnown vendor list CSV: {vendor_csv_path}"
    task += """

Execute the full pipeline: extract, normalize, aggregate, and generate the Excel report. \
Then summarize the results."""

    server = _build_agent_tools()
    options = ClaudeAgentOptions(
        system_prompt=SYSTEM_PROMPT, model=model,
        mcp_servers={"reconciliation": server},
        allowed_tools=[
            "mcp__reconciliation__extract_pdf_transactions",
            "mcp__reconciliation__load_vendor_list",
            "mcp__reconciliation__normalize_vendors",
            "mcp__reconciliation__aggregate_by_vendor",
            "mcp__reconciliation__generate_excel_report",
        ],
        permission_mode="acceptEdits",
    )

    final_text = ""
    total_cost = 0.0
    tool_calls = 0

    async with ClaudeSDKClient(options=options) as client:
        await client.query(task)
        async for message in client.receive_response():
            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        final_text += block.text + "\n"
                    elif hasattr(block, "name"):
                        tool_calls += 1
            if isinstance(message, ResultMessage):
                if hasattr(message, "total_cost_usd") and message.total_cost_usd:
                    total_cost = message.total_cost_usd

    # Pull stats from session state
    summaries = _agent_session_state.get("summaries", [])
    transactions = _agent_session_state.get("transactions", [])
    total_amount = sum(s.total_amount for s in summaries)
    vendors_over_600 = sum(1 for s in summaries if s.total_amount >= 600)
    vendors_needing_review = sum(1 for s in summaries if s.needs_review)

    return {
        "success": True,
        "mode": "agent",
        "file_id": file_id,
        "transaction_count": len(transactions),
        "vendor_count": len(summaries),
        "total_amount": total_amount,
        "vendors_over_600": vendors_over_600,
        "vendors_needing_review": vendors_needing_review,
        "extraction_method": "agent-orchestrated",
        "warnings": [],
        "vendor_preview": [_summary_to_dict(s) for s in summaries],
        "agent_summary": final_text.strip(),
        "agent_cost_usd": total_cost,
        "agent_tool_calls": tool_calls,
        "model": model,
    }
